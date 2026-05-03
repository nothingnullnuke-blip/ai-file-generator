from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from pathlib import Path

from generators.base_generator import BaseGenerator
from models.schemas import DataType
from utils.logger import setup_logger

logger = setup_logger(__name__)

class ExcelGenerator(BaseGenerator):
    def __init__(self, structured_output, output_filename=None):
        super().__init__(structured_output, output_filename)
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Report"
        self.row = 1
    
    def generate(self) -> Path:
        try:
            logger.info(f"Generating Excel: {self.filepath}")
            
            self._add_title()
            
            if self.output.description:
                self._add_description()
            
            for section in sorted(self.output.sections, key=lambda x: x.order):
                self._add_section(section)
            
            self._adjust_columns()
            
            self.wb.save(str(self.filepath))
            logger.info(f"Excel generated: {self.filepath}")
            return self.filepath
            
        except Exception as e:
            logger.error(f"Excel generation failed: {str(e)}")
            raise
    
    def get_file_extension(self) -> str:
        return "xlsx"
    
    def _add_title(self):
        cell = self.ws[f'A{self.row}']
        cell.value = self.output.title
        cell.font = Font(name='Calibri', size=18, bold=True, color='FFFFFF')
        cell.fill = PatternFill(
            start_color=self.colors['primary'].lstrip('#'),
            end_color=self.colors['primary'].lstrip('#'),
            fill_type='solid'
        )
        cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.merge_cells(f'A{self.row}:D{self.row}')
        self.ws.row_dimensions[self.row].height = 30
        self.row += 2
    
    def _add_description(self):
        cell = self.ws[f'A{self.row}']
        cell.value = self.output.description
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        self.ws.merge_cells(f'A{self.row}:D{self.row}')
        self.ws.row_dimensions[self.row].height = 40
        self.row += 2
    
    def _add_section(self, section):
        if section.type == DataType.HEADING:
            self._add_heading(section)
        elif section.type == DataType.TEXT:
            self._add_text(section)
        elif section.type == DataType.TABLE:
            self._add_table(section)
    
    def _add_heading(self, section):
        cell = self.ws[f'A{self.row}']
        cell.value = section.content.get('text', '')
        cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        cell.fill = PatternFill(
            start_color=self.colors['secondary'].lstrip('#'),
            end_color=self.colors['secondary'].lstrip('#'),
            fill_type='solid'
        )
        self.ws.merge_cells(f'A{self.row}:D{self.row}')
        self.ws.row_dimensions[self.row].height = 25
        self.row += 1
    
    def _add_text(self, section):
        cell = self.ws[f'A{self.row}']
        cell.value = section.content.get('text', '')
        cell.alignment = Alignment(wrap_text=True)
        self.ws.merge_cells(f'A{self.row}:D{self.row}')
        self.ws.row_dimensions[self.row].height = 30
        self.row += 2
    
    def _add_table(self, section):
        content = section.content
        title = content.get('title', '')
        headers = content.get('headers', [])
        rows = content.get('rows', [])
        
        if title:
            cell = self.ws[f'A{self.row}']
            cell.value = title
            cell.font = Font(bold=True, size=12)
            self.row += 1
        
        header_fill = PatternFill(
            start_color=self.colors['secondary'].lstrip('#'),
            end_color=self.colors['secondary'].lstrip('#'),
            fill_type='solid'
        )
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self.row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        self.row += 1
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_data in rows:
            for col_idx, value in enumerate(row_data, 1):
                cell = self.ws.cell(row=self.row, column=col_idx)
                cell.value = value
                cell.border = border
            self.row += 1
        
        self.row += 1
    
    def _adjust_columns(self):
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            self.ws.column_dimensions[column_letter].width = min(max_length + 2, 50)